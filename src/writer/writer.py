from __future__ import annotations

import os
from pathlib import Path
from typing import Any, Dict, List, Set

from openpyxl import Workbook, load_workbook
from openpyxl.worksheet.worksheet import Worksheet

from src.ruleresolver.api import RuleSet
from src.extractor.api import AssayRecord
from .model import WriteResult


class WriterError(RuntimeError):
    pass


class Writer:
    def write_record(self, record: AssayRecord, ruleset: RuleSet, output_dir: str) -> WriteResult:
        # Excel-Regeln aus dem Ruleset lesen
        excel_rules = ruleset.data.get("excel_rules", {})
        filename_template = excel_rules.get("excel_filename_template", "{assay_name}.xlsx")
        assay_name = ruleset.data.get("assay_name") or ruleset.assay_key

        sheet_template = excel_rules.get("sheetname_template", "{lot_id}")

        excel_name = filename_template.format(
            assay_key=self._sanitize_filename(ruleset.assay_key),
            assay_name=self._sanitize_filename(assay_name),
        )
        sheet_name = sheet_template.format(lot_id=self._sanitize_sheetname(record.lot_id))

        out_dir = Path(output_dir)
        out_dir.mkdir(parents=True, exist_ok=True)
        excel_path = out_dir / excel_name

        lock_path = out_dir / ".excel_writer.lock"
        self._acquire_lock(lock_path)
        try:
            status = self._write_with_dedupe(excel_path, sheet_name, record, excel_rules)
        finally:
            self._release_lock(lock_path)

        return WriteResult(excel_path=str(excel_path), sheet_name=sheet_name, status=status)

    def _write_with_dedupe(
        self,
        excel_path: Path,
        sheet_name: str,
        record: AssayRecord,
        excel_rules: Dict[str, Any],
    ) -> str:
        if excel_path.exists():
            wb = load_workbook(excel_path)
            status_base = "appended"
        else:
            wb = Workbook()
            # Default-Sheet entfernen
            if "Sheet" in wb.sheetnames and len(wb.sheetnames) == 1:
                wb.remove(wb["Sheet"])
            status_base = "created"

        ws = wb[sheet_name] if sheet_name in wb.sheetnames else wb.create_sheet(sheet_name)

        headers = self._ensure_headers(ws, record, excel_rules)

        # Dedupe prüfen
        existing = self._existing_dedupe_keys(ws, headers)
        if record.dedupe_key in existing:
            wb.save(excel_path)
            return "skipped"

        # Mapping: internal_key -> excel_column_name
        mapping: Dict[str, str] = excel_rules.get("column_mapping", {})
        # Reverse: excel_column_name -> internal_key (fürs Zurückübersetzen beim Schreiben)
        reverse_mapping: Dict[str, str] = {v: k for k, v in mapping.items()}

        row_values: List[Any] = []
        for h in headers:
            if h == "assay_key":
                row_values.append(record.assay_key)
            elif h == "lot_id":
                row_values.append(record.lot_id)
            elif h == "dedupe_key":
                row_values.append(record.dedupe_key)
            else:
                # Wichtig: Wenn Header gemappt ist (z.B. "Haltbarkeit"), dann den originalen Key nehmen (z.B. "expiry_raw")
                internal_key = reverse_mapping.get(h, h)
                row_values.append(record.data.get(internal_key))

        ws.append(row_values)

        wb.save(excel_path)
        return status_base

    def _existing_dedupe_keys(self, ws: Worksheet, headers: List[str]) -> Set[str]:
        if "dedupe_key" not in headers:
            return set()
        dedupe_col = headers.index("dedupe_key") + 1
        keys: Set[str] = set()
        for row in ws.iter_rows(min_row=2, values_only=True):
            if row and len(row) >= dedupe_col:
                v = row[dedupe_col - 1]
                if v:
                    keys.add(str(v))
        return keys

    def _ensure_headers(self, ws: Worksheet, record: AssayRecord, excel_rules: Dict[str, Any]) -> List[str]:
        mapping: Dict[str, str] = excel_rules.get("column_mapping", {})
        base_cols = ["assay_key", "lot_id", "dedupe_key"]

        # Excel-Header-Namen (gemappt), aber Keys bleiben intern im record.data
        data_cols = [mapping.get(k, k) for k in record.data.keys()]
        desired = base_cols + data_cols

        # Create header row if empty
        if ws.max_row < 1 or ws.cell(1, 1).value is None:
            ws.append(desired)
            return desired

        existing = [c.value for c in ws[1] if c.value is not None]
        if not existing:
            ws.append(desired)
            return desired

        # Append missing columns (keeps existing order)
        existing_set = set(existing)
        for h in desired:
            if h not in existing_set:
                existing.append(h)
                ws.cell(row=1, column=len(existing)).value = h

        return existing

    def _sanitize_filename(self, s: str) -> str:
        return "".join(ch for ch in str(s) if ch.isalnum() or ch in (" ", "_", "-", ".")).strip().replace(" ", "_")

    def _sanitize_sheetname(self, s: str) -> str:
        invalid = set(':\\/?"*[]')
        cleaned = "".join(ch for ch in str(s) if ch not in invalid).strip()
        return cleaned[:31] if cleaned else "LOT"

    def _acquire_lock(self, lock_path: Path) -> None:
        try:
            fd = os.open(str(lock_path), os.O_CREAT | os.O_EXCL | os.O_WRONLY)
            os.close(fd)
        except FileExistsError:
            raise WriterError("excel_writer_lock_exists")

    def _release_lock(self, lock_path: Path) -> None:
        try:
            lock_path.unlink(missing_ok=True)
        except Exception:
            pass
